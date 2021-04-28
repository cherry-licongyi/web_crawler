# -*- coding: utf-8 -*-
#!/usr/bin/env Python 3.8.5
# Author: cherry-licongyi
""" 从excel表中读取视频源url并下载保存视频 """
import requests
from openpyxl import load_workbook
import os

def video_download(fileName,folderName):
    # 请求头信息
    myheader = {'User-Agent':'Mozila/5.0 (Windows NT 10.0;WOW64) AppleWebKit/537.36(KHTML,like Gecko) Chrome/83.0.4103.61 Safari/537.36'}
    excel=load_workbook(fileName)
    table = excel['Sheet']   #通过表名获取sheet 
    rows=table.max_row      #获取行数
    cols=table.max_column    #获取列数
    j = 0
    for i in range(rows): # 下载3个视频
        #获取表格内容，是从第一行第一列是从1开始的，第一行1为表格标签，从2开始获取值
        title = table.cell(row=i+2,column=2).value #获取标题
        title = title.replace(' ','')
        video_url = table.cell(row=i+2,column=4).value   # url
        if video_url == None:
            continue
        response = requests.get(url=video_url,headers=myheader)
        if response.status_code == 200:
            data = response.content
            file = open(folderName + title +'.mp4','wb')
            file.write(data)
            file.close()
            j += 1
            if j == 3:
                break


if __name__ == '__main__':
    menu = ['ifeng','baidu','thepaper','haokan','meipai','cntv','tudou']
    for i in menu:       
        isExists=os.path.exists(i)  # 判断文件夹是否存在        
        if not isExists:             # 不存则创建文件夹
            os.makedirs(i) 
        if i == 'cntv':
            break      
        print('Downloading videos from ' + i )
        filename = i + '.xlsx'
        foldname = i + '/'        
        video_download(fileName=filename,folderName=foldname) # 依次下载视频
    print('\nDone')

